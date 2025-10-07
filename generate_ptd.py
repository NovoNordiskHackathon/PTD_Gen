#!/usr/bin/env python3
"""
Generate a combined PTD workbook with two sheets:
 - Sheet 1: Schedule Grid (from protocol + eCRF)
 - Sheet 2: Study Specific Forms (from eCRF)

The computation logic is reused from existing modules and scripts; this file
organizes them into a single configurable CLI entrypoint and merges outputs.

Modifications:
 - Added --template argument to load an existing template workbook
 - Generate Schedule Grid and Study Specific Forms into temporary files
 - Replace sheets named "Schedule Grid" and "Study Specific Forms" in the template
   (including styles, merges, and dimensions), preserve other sheets, and save to --out
 - Removed the old append/merge flow that built a new workbook from scratch
"""

import os
import sys
import json
import logging
import argparse
from typing import Dict, Any, Optional, List
from pathlib import Path
import tempfile
import shutil
from openpyxl import load_workbook
from openpyxl.worksheet.worksheet import Worksheet
from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.cell.cell import MergedCell

# Reuse existing modules for schedule grid pipeline
from modules.form_extractor import extract_forms
from modules.soa_parser import parse_soa
from modules.common_matrix import merge_common_matrix
from modules.event_grouping import group_events
from modules.schedule_layout import generate_schedule_grid as build_schedule_grid_file


def load_json(file_path: str) -> Dict[str, Any]:
    with open(file_path, "r", encoding="utf-8") as f:
        return json.load(f)


def load_config(config_path: str) -> Dict[str, Any]:
    try:
        with open(config_path, 'r', encoding='utf-8') as f:
            return json.load(f)
    except FileNotFoundError:
        logging.warning(f"Config file not found: {config_path}; using defaults")
        return {}
    except json.JSONDecodeError as e:
        logging.warning(f"Invalid JSON in config {config_path}: {e}; using defaults")
        return {}


def setup_logging(level: str = "INFO") -> None:
    logging.basicConfig(
        level=getattr(logging, level.upper()),
        format="%(asctime)s - %(levelname)s - %(message)s",
        handlers=[
            logging.StreamHandler(sys.stdout),
            logging.FileHandler("ptd_generation.log")
        ]
    )


def ensure_output_dir(output_path: str) -> None:
    out_dir = os.path.dirname(output_path)
    if out_dir:
        Path(out_dir).mkdir(parents=True, exist_ok=True)


def run_schedule_grid_pipeline(protocol_json: str, ecrf_json: str, final_output_xlsx: str, config_dir: str) -> str:
    """
    Reuse the existing 5-stage pipeline to produce the schedule grid Excel file directly
    at final_output_xlsx. Returns the absolute path to the generated file.
    """
    config_files = {
        'form_extractor': 'config_form_extractor.json',
        'soa_parser': 'config_soa_parser.json',
        'common_matrix': 'config_common_matrix.json',
        'event_grouping': 'config_event_grouping.json',
        'schedule_layout': 'config_schedule_layout.json'
    }

    configs: Dict[str, Any] = {}
    for key, filename in config_files.items():
        configs[key] = load_config(os.path.join(config_dir, filename))

    temp_dir = tempfile.mkdtemp(prefix="ptd_intermediate_")
    intermediates: List[str] = []
    try:
        forms_csv = os.path.join(temp_dir, "extracted_forms.csv")
        extract_forms(ecrf_json=ecrf_json, output_csv=forms_csv, config=configs.get('form_extractor', {}))
        intermediates.append(forms_csv)

        schedule_csv = os.path.join(temp_dir, "schedule.csv")
        parse_soa(protocol_json=protocol_json, output_csv=schedule_csv, config=configs.get('soa_parser', {}))
        intermediates.append(schedule_csv)

        matrix_csv = os.path.join(temp_dir, "soa_matrix.csv")
        merge_common_matrix(ecrf_csv=forms_csv, schedule_csv=schedule_csv, output_csv=matrix_csv, config=configs.get('common_matrix', {}))
        intermediates.append(matrix_csv)

        visits_xlsx = os.path.join(temp_dir, "visits_with_groups.xlsx")
        group_events(protocol_json=protocol_json, output_xlsx=visits_xlsx, config=configs.get('event_grouping', {}))
        intermediates.append(visits_xlsx)

        ensure_output_dir(final_output_xlsx)
        build_schedule_grid_file(visits_xlsx=visits_xlsx, forms_csv=matrix_csv, output_xlsx=final_output_xlsx, config=configs.get('schedule_layout', {}))

        return os.path.abspath(final_output_xlsx)
    finally:
        # Best-effort cleanup
        for p in intermediates:
            try:
                if os.path.exists(p):
                    os.remove(p)
            except Exception:
                pass
        try:
            shutil.rmtree(temp_dir)
        except Exception:
            pass


def generate_study_specific_forms_xlsx(ecrf_json: str) -> str:
    """
    Reuse logic from Final_study_specific_form.py by invoking its processing function to
    produce an Excel file. Returns the path to the generated temp Excel.
    """
    # Import here to avoid executing module-level code unless needed
    import importlib.util

    module_path = os.path.join(os.path.dirname(__file__), 'Final_study_specific_form.py')
    spec = importlib.util.spec_from_file_location("Final_study_specific_form", module_path)
    if spec is None or spec.loader is None:
        raise RuntimeError("Unable to load Final_study_specific_form.py")
    mod = importlib.util.module_from_spec(spec)
    spec.loader.exec_module(mod)

    temp_dir = tempfile.mkdtemp(prefix="ptd_forms_")
    output_xlsx = os.path.join(temp_dir, "study_specific_forms.xlsx")

    # The script's API function writes the Excel; keep its computation logic intact
    config_rules = os.path.join(os.path.dirname(__file__), 'config', 'config_study_specific_forms.json')
    # Avoid hardcoded/unnecessary template path; rely on the module's internal template
    mod.process_clinical_forms(ecrf_json, output_csv_path=output_xlsx, config_path=config_rules)
    return output_xlsx


## Removed: unused header renaming/ordering helper.


def _copy_worksheet_contents(src_ws: Worksheet, dest_ws: Worksheet) -> None:
    """Copy values, styles, merged cells, and dimensions from src_ws to dest_ws."""
    # Copy column widths
    for col_letter, dim in src_ws.column_dimensions.items():
        if getattr(dim, 'width', None):
            dest_ws.column_dimensions[col_letter].width = dim.width

    # Copy row heights
    for idx, dim in src_ws.row_dimensions.items():
        if getattr(dim, 'height', None):
            dest_ws.row_dimensions[idx].height = dim.height

    # Copy merged ranges first (structure)
    for merged_range in src_ws.merged_cells.ranges:
        dest_ws.merge_cells(str(merged_range))

    # Copy cell contents and styles
    for row in src_ws.iter_rows():
        for cell in row:
            # Skip non-top-left merged cells; the range has already been created
            if isinstance(cell, MergedCell):
                continue
            dcell = dest_ws.cell(row=cell.row, column=cell.column, value=cell.value)
            if cell.has_style:
                if cell.font:
                    dcell.font = Font(
                        name=cell.font.name,
                        size=cell.font.size,
                        bold=cell.font.bold,
                        italic=cell.font.italic,
                        vertAlign=cell.font.vertAlign,
                        underline=cell.font.underline,
                        strike=cell.font.strike,
                        color=cell.font.color,
                    )
                if cell.alignment:
                    dcell.alignment = Alignment(
                        horizontal=cell.alignment.horizontal,
                        vertical=cell.alignment.vertical,
                        text_rotation=cell.alignment.text_rotation,
                        wrap_text=cell.alignment.wrap_text,
                        shrink_to_fit=cell.alignment.shrink_to_fit,
                        indent=cell.alignment.indent,
                    )
                if cell.fill and cell.fill.fill_type:
                    dcell.fill = PatternFill(
                        fill_type=cell.fill.fill_type,
                        start_color=cell.fill.start_color,
                        end_color=cell.fill.end_color,
                    )
                if cell.border:
                    left = cell.border.left
                    right = cell.border.right
                    top = cell.border.top
                    bottom = cell.border.bottom
                    dcell.border = Border(
                        left=Side(style=left.style, color=left.color),
                        right=Side(style=right.style, color=right.color),
                        top=Side(style=top.style, color=top.color),
                        bottom=Side(style=bottom.style, color=bottom.color),
                    )
                if cell.number_format:
                    dcell.number_format = cell.number_format


def _copy_worksheet_values_only(src_ws: Worksheet, dest_ws: Worksheet) -> None:
    """
    Fast path: copy only cell values row-by-row using append.
    Skips styles, merges, widths, heights for speed.
    """
    for row in src_ws.iter_rows(values_only=True):
        dest_ws.append(list(row))

    # Preserve merged cell structure if available (ReadOnlyWorksheet may not expose it)
    try:
        merged_ranges = None
        merged_container = getattr(src_ws, "merged_cells", None)
        if merged_container is not None:
            merged_ranges = getattr(merged_container, "ranges", None)
        if not merged_ranges:
            # Fallback for older openpyxl
            merged_ranges = getattr(src_ws, "merged_cell_ranges", None)
        if merged_ranges:
            for merged_range in merged_ranges:
                try:
                    dest_ws.merge_cells(str(merged_range))
                except Exception:
                    pass
    except Exception:
        # If unavailable in read-only mode, skip merges for speed/simplicity
        pass


def _copy_header_styles(src_ws: Worksheet, dest_ws: Worksheet, max_header_rows: int) -> None:
    """
    Copy only styles (font, alignment, fill, border, number_format) for the first
    max_header_rows to preserve header/group colouring while keeping operation fast.
    """
    if max_header_rows <= 0:
        return
    rows_to_copy = min(max_header_rows, src_ws.max_row)
    for r in range(1, rows_to_copy + 1):
        for cell in src_ws[r]:
            try:
                dcell = dest_ws.cell(row=cell.row, column=cell.column)
                if cell.has_style:
                    if cell.font:
                        dcell.font = Font(
                            name=cell.font.name,
                            size=cell.font.size,
                            bold=cell.font.bold,
                            italic=cell.font.italic,
                            vertAlign=cell.font.vertAlign,
                            underline=cell.font.underline,
                            strike=cell.font.strike,
                            color=cell.font.color,
                        )
                    if cell.alignment:
                        dcell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical=cell.alignment.vertical,
                            text_rotation=cell.alignment.text_rotation,
                            wrap_text=cell.alignment.wrap_text,
                            shrink_to_fit=cell.alignment.shrink_to_fit,
                            indent=cell.alignment.indent,
                        )
                    if cell.fill and cell.fill.fill_type:
                        dcell.fill = PatternFill(
                            fill_type=cell.fill.fill_type,
                            start_color=cell.fill.start_color,
                            end_color=cell.fill.end_color,
                        )
                    if cell.border:
                        left = cell.border.left
                        right = cell.border.right
                        top = cell.border.top
                        bottom = cell.border.bottom
                        dcell.border = Border(
                            left=Side(style=left.style, color=left.color),
                            right=Side(style=right.style, color=right.color),
                            top=Side(style=top.style, color=top.color),
                            bottom=Side(style=bottom.style, color=bottom.color),
                        )
                    if cell.number_format:
                        dcell.number_format = cell.number_format
            except Exception:
                # Best-effort; continue
                pass


def replace_sheets_in_template(
    template_xlsx: str,
    schedule_xlsx: str,
    forms_xlsx: str,
    out_xlsx: str,
    schedule_sheet_name: str = "Schedule Grid",
    forms_sheet_name: str = "Study Specific Forms",
    fast: bool = False,
) -> str:
    """
    Load the template workbook, remove existing target sheets if present, copy
    the generated schedule and forms worksheets (including styles, merges, and
    dimensions) into the template, preserve all other sheets, and save to out_xlsx.
    Returns the absolute path to the saved workbook.
    """
    ensure_output_dir(out_xlsx)

    wb_template = load_workbook(template_xlsx)
    # Load sources; for fast path we still load normally to allow header style copy
    wb_schedule = load_workbook(schedule_xlsx)
    wb_forms = load_workbook(forms_xlsx)

    try:
        # Determine insertion indices to preserve original order if sheets existed
        schedule_index = None
        forms_index = None
        if schedule_sheet_name in wb_template.sheetnames:
            schedule_index = wb_template.sheetnames.index(schedule_sheet_name)
            wb_template.remove(wb_template[schedule_sheet_name])
        if forms_sheet_name in wb_template.sheetnames:
            forms_index = wb_template.sheetnames.index(forms_sheet_name)
            wb_template.remove(wb_template[forms_sheet_name])

        # Create destination sheets at recorded positions (or append if None)
        if schedule_index is not None:
            dest_schedule = wb_template.create_sheet(title=schedule_sheet_name, index=schedule_index)
        else:
            dest_schedule = wb_template.create_sheet(title=schedule_sheet_name)
        if forms_index is not None:
            dest_forms = wb_template.create_sheet(title=forms_sheet_name, index=forms_index)
        else:
            dest_forms = wb_template.create_sheet(title=forms_sheet_name)

        # Source sheets (first worksheet in each generated file)
        src_schedule: Worksheet = wb_schedule.worksheets[0]
        src_forms: Worksheet = wb_forms.worksheets[0]

        # Copy contents
        if fast:
            _copy_worksheet_values_only(src_schedule, dest_schedule)
            _copy_worksheet_values_only(src_forms, dest_forms)
            # Preserve header styles (keep column/group colouring)
            _copy_header_styles(src_schedule, dest_schedule, max_header_rows=5)
            _copy_header_styles(src_forms, dest_forms, max_header_rows=3)
        else:
            _copy_worksheet_contents(src_schedule, dest_schedule)
            _copy_worksheet_contents(src_forms, dest_forms)

        wb_template.save(out_xlsx)
        return os.path.abspath(out_xlsx)
    finally:
        try:
            wb_schedule.close()
        except Exception:
            pass
        try:
            wb_forms.close()
        except Exception:
            pass
        try:
            wb_template.close()
        except Exception:
            pass


def auto_format_sheet(sheet: Worksheet, header_rows: int = 1, skip_fill_rows=None) -> None:
    """Auto-fit columns, style headers (one or more rows), and apply borders.

    Preserves any existing header fills and avoids filling Row 1 beyond column D.
    """
    header_font = Font(bold=True)
    center_align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    header_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
    thin_border = Border(
        left=Side(style="thin"), right=Side(style="thin"),
        top=Side(style="thin"), bottom=Side(style="thin")
    )

    # Style header rows
    header_rows = max(1, min(header_rows, sheet.max_row))
    skip_fill_rows = skip_fill_rows or set()
    for r in range(1, header_rows + 1):
        for cell in sheet[r]:
            cell.font = header_font
            cell.alignment = center_align
            # Preserve pre-existing fills (do not override group colors)
            try:
                fill_type = getattr(cell.fill, 'fill_type', None)
            except Exception:
                fill_type = None
            # Do not apply header fill to Row 1 columns beyond D
            col_idx = cell.column if hasattr(cell, 'column') else cell.col_idx
            beyond_ctdm = (r == 1 and col_idx and col_idx > 4)
            if not fill_type and r not in skip_fill_rows and not beyond_ctdm:
                cell.fill = header_fill
            cell.border = thin_border

    # Style all data rows
    for row in sheet.iter_rows(min_row=header_rows + 1, max_row=sheet.max_row, max_col=sheet.max_column):
        for cell in row:
            cell.alignment = Alignment(wrap_text=True, vertical="center")
            cell.border = thin_border

    # Auto-adjust column widths
    for col in sheet.columns:
        max_length = 0
        first_cell = next(iter(col))
        column_letter = get_column_letter(first_cell.column)
        for cell in col:
            try:
                val_len = len(str(cell.value)) if cell.value is not None else 0
                if val_len > max_length:
                    max_length = val_len
            except Exception:
                pass
        sheet.column_dimensions[column_letter].width = max(10, min(80, max_length + 3))


def finalize_formatting(output_path: str, forms_sheet_name: str = "Study Specific Forms") -> None:
    """
    Apply formatting only to the study-specific forms sheet. The schedule grid
    formatting produced by its generator is preserved as-is.
    """
    wb = load_workbook(output_path)
    if forms_sheet_name in wb.sheetnames:
        ws = wb[forms_sheet_name]
        # Keep the 3 fixed header rows (Row 1 CTDM, Row 2 merged groups, Row 3 subheaders)
        auto_format_sheet(ws, header_rows=3, skip_fill_rows={})
    wb.save(output_path)


def main() -> int:
    parser = argparse.ArgumentParser(
        description="Generate PTD Excel with Schedule Grid and Study Specific Forms"
    )
    parser.add_argument("--ecrf", required=True, help="Path to hierarchical_output_final_ecrf.json")
    parser.add_argument("--protocol", required=True, help="Path to hierarchical_output_final_protocol.json")
    parser.add_argument("--template", required=True, help="Path to template Excel (will be updated)")
    parser.add_argument("--out", required=False, help="Output Excel file path (e.g., ptd.xlsx). Omit when using --inplace")
    parser.add_argument("--inplace", action="store_true", help="Modify the template file in place (save over --template)")
    parser.add_argument("--fast", action="store_true", help="Fast mode: values-only copy, skip extra formatting")
    args = parser.parse_args()

    setup_logging("INFO")

    # Determine output path (in-place or new file)
    if args.inplace:
        output_path = args.template
        # Warn if --out was also provided but different
        if args.out and os.path.abspath(args.out) != os.path.abspath(args.template):
            logging.warning("--inplace specified: ignoring --out and writing to template path")
    else:
        if not args.out:
            print("Error: --out is required unless --inplace is specified", file=sys.stderr)
            return 2
        output_path = args.out

    # Normalize output path extension and ensure folder
    if not output_path.lower().endswith(".xlsx"):
        output_path = os.path.splitext(output_path)[0] + ".xlsx"
    ensure_output_dir(output_path)

    # 1) Build schedule grid into a temp workbook
    schedule_tmp_dir = tempfile.mkdtemp(prefix="ptd_schedule_")
    schedule_tmp_xlsx = os.path.join(schedule_tmp_dir, "schedule_grid.xlsx")
    schedule_path = run_schedule_grid_pipeline(
        protocol_json=args.protocol,
        ecrf_json=args.ecrf,
        final_output_xlsx=schedule_tmp_xlsx,
        config_dir=os.path.join(os.path.dirname(__file__), "config"),
    )

    # 2) Generate study specific forms to a temp file
    forms_tmp_xlsx = generate_study_specific_forms_xlsx(args.ecrf)

    # 3) Replace sheets in the provided template and save to output
    final_path = replace_sheets_in_template(
        template_xlsx=args.template,
        schedule_xlsx=schedule_path,
        forms_xlsx=forms_tmp_xlsx,
        out_xlsx=output_path,
        fast=args.fast,
    )

    # 4) Finalize formatting on forms sheet (skip in fast mode)
    if not args.fast:
        finalize_formatting(final_path)

    # Cleanup temp dirs
    try:
        shutil.rmtree(schedule_tmp_dir)
    except Exception:
        pass
    try:
        shutil.rmtree(os.path.dirname(forms_tmp_xlsx))
    except Exception:
        pass

    print(f"✅ Combined PTD file written successfully to: {final_path}")
    return 0


if __name__ == "__main__":
    sys.exit(main())